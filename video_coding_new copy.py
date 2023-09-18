import pandas as pd
import os
import cv2
import openpyxl
import numpy as np
import matplotlib.pyplot as plt
from tkinter import *
from tkinter import ttk
from PIL import ImageTk, Image

event_summary = pd.read_csv("Events/Pedestrian/NCRC_Events-1.csv")

class coding_GUI():
    def __init__(self, init_window_name):
        self.init_window_name = init_window_name
        self.init_window_name.title("Mcity video coding")
        self.init_window_name.geometry('1920x1080+0+0')

        # canvas
        self.canvas = Canvas(width=1280, height=720)
        self.canvas.place(x=10,y=10)

        # load event
        self.event_num = Label(self.init_window_name, text="event num")
        self.event_num.place(x=-500+1310,y=10)
        self.event_num.config(font=(20))

        self.event_num_entry = Entry(self.init_window_name, font=(20), width=15)
        self.event_num_entry.place(x=-500+1400,y=10)
        self.event_num_entry.bind("<Return>", self.load_video)
        self.event_num_entry.bind("<Left>", self.prev_second)
        self.event_num_entry.bind("<Right>", self.next_second)

        self.load_video_button = Button(self.init_window_name, text="load", width=15, bg='light blue', command=self.load_video)
        self.load_video_button.place(x=-500+1560,y=5)
        self.load_video_button.config(font=(20))

        # progress checking
        self.complete = Text(self.init_window_name, height=1, width=30)
        self.complete.place(x=-500+1310,y=40)
        self.complete.config(font=(20))

        # time button
        self.prev_frame_button = Button(self.init_window_name, text="<<", width=6, command=self.prev_frame)
        self.prev_frame_button.place(x=-500+1410, y=80)

        self.next_frame_button = Button(self.init_window_name, text=">>", width=6, command=self.next_frame)
        self.next_frame_button.place(x=-500+1510,y=80)

        self.prev_second_button = Button(self.init_window_name, text="<<<", width=6, command=self.prev_second)
        self.prev_second_button.place(x=-500+1310,y=80)

        self.next_second_button = Button(self.init_window_name, text=">>>", width=6, command=self.next_second)
        self.next_second_button.place(x=-500+1610,y=80)

        # right of way
        self.right_of_way = Label(self.init_window_name, text="1. right of way")
        self.right_of_way.place(x=-500+1310,y=120)
        self.right_of_way.config(font=(20))

        self.right_of_way_listbox = ttk.Combobox(self.init_window_name, values=["Shuttle Bus", "Pedestrian", "Others"])
        self.right_of_way_listbox.place(x=-500+1460,y=120)
        self.right_of_way_listbox.config(font=(20))
        self.right_of_way_listbox.bind("<<ComboboxSelected>>", self.right_of_way_callback)

        self.right_of_way_note = Label(self.init_window_name, text="note:")
        self.right_of_way_note.place(x=-500+1680,y=120)
        self.right_of_way_note.config(font=(20))

        self.right_of_way_note_entry = Entry(self.init_window_name, font=(20), width=25)
        self.right_of_way_note_entry.place(x=-500+1730,y=120)
        self.right_of_way_note_entry.bind("<Return>", self.right_of_way_note_entry_callback)

        # secondary activity
        self.secondary_activity = Label(self.init_window_name, text="2. secondary activity")
        self.secondary_activity.place(x=-500+1310,y=160)
        self.secondary_activity.config(font=(20))

        # None
        self.checkvalue_1 = BooleanVar() 
        self.sec_checkbutton_1 = Checkbutton(self.init_window_name, text='None', var=self.checkvalue_1, command=self.checkbutton_1_callback)
        self.sec_checkbutton_1.place(x=-500+1310,y=180)
        self.sec_checkbutton_1.config(font=(20))

        # Phone, listening or talking
        self.checkvalue_2 = BooleanVar() 
        self.sec_checkbutton_2 = Checkbutton(self.init_window_name, text='Phone, listening or talking', var=self.checkvalue_2, command=self.checkbutton_2_callback)
        self.sec_checkbutton_2.place(x=-500+1310,y=210)
        self.sec_checkbutton_2.config(font=(20))

        # Phone, looking and/or manipulating
        self.checkvalue_3 = BooleanVar() 
        self.sec_checkbutton_3 = Checkbutton(self.init_window_name, text='Phone, looking and/or manipulating', var=self.checkvalue_3, command=self.checkbutton_3_callback)
        self.sec_checkbutton_3.place(x=-500+1310,y=240)
        self.sec_checkbutton_3.config(font=(20))

        # Looking at personal item
        self.checkvalue_4 = BooleanVar() 
        self.sec_checkbutton_4 = Checkbutton(self.init_window_name, text='Looking at personal item', var=self.checkvalue_4, command=self.checkbutton_4_callback)
        self.sec_checkbutton_4.place(x=-500+1310,y=270)
        self.sec_checkbutton_4.config(font=(20))

        # Talking to someone in person
        self.checkvalue_5 = BooleanVar() 
        self.sec_checkbutton_5 = Checkbutton(self.init_window_name, text='Talking to someone in person', var=self.checkvalue_5, command=self.checkbutton_5_callback)
        self.sec_checkbutton_5.place(x=-500+1310,y=300)
        self.sec_checkbutton_5.config(font=(20))

        # Gesturing or interacting
        self.checkvalue_6 = BooleanVar() 
        self.sec_checkbutton_6 = Checkbutton(self.init_window_name, text='Gesturing or interacting', var=self.checkvalue_6, command=self.checkbutton_6_callback)
        self.sec_checkbutton_6.place(x=-500+1310,y=330)
        self.sec_checkbutton_6.config(font=(20))

        # Eating or drinking
        self.checkvalue_7 = BooleanVar() 
        self.sec_checkbutton_7 = Checkbutton(self.init_window_name, text='Eating or drinking', var=self.checkvalue_7, command=self.checkbutton_7_callback)
        self.sec_checkbutton_7.place(x=-500+1310,y=360)
        self.sec_checkbutton_7.config(font=(20))

        # Jogging, recreation
        self.checkvalue_8 = BooleanVar() 
        self.sec_checkbutton_8 = Checkbutton(self.init_window_name, text='Jogging, recreation', var=self.checkvalue_8, command=self.checkbutton_8_callback)
        self.sec_checkbutton_8.place(x=-500+1680,y=180)
        self.sec_checkbutton_8.config(font=(20))

        # Apparent rushing
        self.checkvalue_9 = BooleanVar() 
        self.sec_checkbutton_9 = Checkbutton(self.init_window_name, text='Apparent rushing', var=self.checkvalue_9, command=self.checkbutton_9_callback)
        self.sec_checkbutton_9.place(x=-500+1680,y=210)
        self.sec_checkbutton_9.config(font=(20))

        # Earphones
        self.checkvalue_10 = BooleanVar() 
        self.sec_checkbutton_10 = Checkbutton(self.init_window_name, text='Earphones', var=self.checkvalue_10, command=self.checkbutton_10_callback)
        self.sec_checkbutton_10.place(x=-500+1680,y=240)
        self.sec_checkbutton_10.config(font=(20))

        # Other
        self.checkvalue_11 = BooleanVar() 
        self.sec_checkbutton_11 = Checkbutton(self.init_window_name, text='Other', var=self.checkvalue_11, command=self.checkbutton_11_callback)
        self.sec_checkbutton_11.place(x=-500+1680,y=270)
        self.sec_checkbutton_11.config(font=(20))

        # Earphones
        self.checkvalue_12 = BooleanVar() 
        self.sec_checkbutton_12 = Checkbutton(self.init_window_name, text='Unable to determine', var=self.checkvalue_12, command=self.checkbutton_12_callback)
        self.sec_checkbutton_12.place(x=-500+1680,y=300)
        self.sec_checkbutton_12.config(font=(20))

        self.secondary_activity_note = Label(self.init_window_name, text="note:")
        self.secondary_activity_note.place(x=-500+1680,y=160)
        self.secondary_activity_note.config(font=(20))

        self.secondary_activity_note_entry = Entry(self.init_window_name, font=(20), width=25)
        self.secondary_activity_note_entry.place(x=-500+1730,y=160)
        self.secondary_activity_note_entry.bind("<Return>", self.secondary_activity_note_entry_callback)

        # First road user glance to study vehicle
        self.glance = Label(self.init_window_name, text="3. first road user glance to study vehicle")
        self.glance.place(x=-500+1310,y=400)
        self.glance.config(font=(20))

        self.glance_button = Button(self.init_window_name, text="yes", width=6, command=self.glance_yes)
        self.glance_button.place(x=-500+1590,y=395)
        self.glance_button.config(font=(20))

        self.glance_note = Label(self.init_window_name, text="note:")
        self.glance_note.place(x=-500+1680,y=400)
        self.glance_note.config(font=(20))

        self.glance_note_entry = Entry(self.init_window_name, font=(20), width=25)
        self.glance_note_entry.place(x=-500+1730,y=400)
        self.glance_note_entry.bind("<Return>", self.glance_note_entry_callback)

        # Body languages
        self.hand = Label(self.init_window_name, text="4. body languages")
        self.hand.place(x=-500+1310,y=440)
        self.hand.config(font=(20))

        self.hand_listbox = ttk.Combobox(self.init_window_name, values=["Hand gesture to let the shuttle go", "Hand gesture to stop the shuttle", "Others"])
        self.hand_listbox.place(x=-500+1460,y=440)
        self.hand_listbox.config(font=(20))
        self.hand_listbox.bind("<<ComboboxSelected>>", self.hand_callback)

        self.hand_note = Label(self.init_window_name, text="note:")
        self.hand_note.place(x=-500+1680,y=440)
        self.hand_note.config(font=(20))

        self.hand_note_entry = Entry(self.init_window_name, font=(20), width=25)
        self.hand_note_entry.place(x=-500+1730,y=440)
        self.hand_note_entry.bind("<Return>", self.hand_note_entry_callback)

        # ChangeT2
        self.change_t2 = Label(self.init_window_name, text="5. changeT2")
        self.change_t2.place(x=-500+1310,y=480)
        self.change_t2.config(font=(20))

        self.change_t2_button = Button(self.init_window_name, text="yes", width=6, command=self.change_t2_yes)
        self.change_t2_button.place(x=-500+1590,y=475)
        self.change_t2_button.config(font=(20))

        self.change_t2_note = Label(self.init_window_name, text="note:")
        self.change_t2_note.place(x=-500+1680,y=480)
        self.change_t2_note.config(font=(20))

        self.change_t2_note_entry = Entry(self.init_window_name, font=(20), width=25)
        self.change_t2_note_entry.place(x=-500+1730,y=480)
        self.change_t2_note_entry.bind("<Return>", self.change_t2_note_entry_callback)

        # reactionT2
        self.reaction_t2 = Label(self.init_window_name, text="6. reactionT2")
        self.reaction_t2.place(x=-500+1310,y=520)
        self.reaction_t2.config(font=(20))

        self.reaction_t2_listbox = ttk.Combobox(self.init_window_name, values=["Begin walking from stationary", "Continue walking but accelerate or change path", "Continue walking but decelerate", "Interrupted walking then continue", "Interrupted walking and abort", "Other", "Not Applicable", "Unable to determine"])
        self.reaction_t2_listbox.place(x=-500+1460,y=520)
        self.reaction_t2_listbox.config(font=(20))
        self.reaction_t2_listbox.bind("<<ComboboxSelected>>", self.reaction_t2_callback)

        self.reaction_t2_note = Label(self.init_window_name, text="note:")
        self.reaction_t2_note.place(x=-500+1680,y=520)
        self.reaction_t2_note.config(font=(20))

        self.reaction_t2_note_entry = Entry(self.init_window_name, font=(20), width=25)
        self.reaction_t2_note_entry.place(x=-500+1730,y=520)
        self.reaction_t2_note_entry.bind("<Return>", self.reaction_t2_note_entry_callback)

        # block
        self.block = Label(self.init_window_name, text="7. block")
        self.block.place(x=-500+1310,y=560)
        self.block.config(font=(20))

        self.block_listbox = ttk.Combobox(self.init_window_name, values=["Yes", "No", "Other"])
        self.block_listbox.place(x=-500+1460,y=560)
        self.block_listbox.config(font=(20))
        self.block_listbox.bind("<<ComboboxSelected>>", self.block_callback)

        self.block_note = Label(self.init_window_name, text="note:")
        self.block_note.place(x=-500+1680,y=560)
        self.block_note.config(font=(20))

        self.block_note_entry = Entry(self.init_window_name, font=(20), width=25)
        self.block_note_entry.place(x=-500+1730,y=560)
        self.block_note_entry.bind("<Return>", self.block_note_entry_callback)

        # weather 
        self.weather = Label(self.init_window_name, text="8. weather")
        self.weather.place(x=-500+1310,y=640)
        self.weather.config(font=(20))

        self.weather_listbox = ttk.Combobox(self.init_window_name, values=["Sunny/partly cloudy", "Overcast", "Rain/snow/fog/mist", "Unable to determine"])
        self.weather_listbox.place(x=-500+1460,y=640)
        self.weather_listbox.config(font=(20))
        self.weather_listbox.bind("<<ComboboxSelected>>", self.weather_callback)

        self.weather_note = Label(self.init_window_name, text="note:")
        self.weather_note.place(x=-500+1680,y=640)
        self.weather_note.config(font=(20))

        self.weather_note_entry = Entry(self.init_window_name, font=(20), width=25)
        self.weather_note_entry.place(x=-500+1730,y=640)
        self.weather_note_entry.bind("<Return>", self.weather_note_entry_callback)

        # Age group
        self.age_group = Label(self.init_window_name, text="9. age group")
        self.age_group.place(x=-500+1310,y=680)
        self.age_group.config(font=(20))

        self.age_group_listbox = ttk.Combobox(self.init_window_name, values=["Younger", "Middle", "Older", "Unable to determine"])
        self.age_group_listbox.place(x=-500+1460,y=680)
        self.age_group_listbox.config(font=(20))
        self.age_group_listbox.bind("<<ComboboxSelected>>", self.age_group_callback)

        self.age_group_note = Label(self.init_window_name, text="note:")
        self.age_group_note.place(x=-500+1680,y=680)
        self.age_group_note.config(font=(20))

        self.age_group_note_entry = Entry(self.init_window_name, font=(20), width=25)
        self.age_group_note_entry.place(x=-500+1730,y=680)
        self.age_group_note_entry.bind("<Return>", self.age_group_note_entry_callback)

        # Gender
        self.gender = Label(self.init_window_name, text="10. gender")
        self.gender.place(x=-500+1310,y=720)
        self.gender.config(font=(20))

        self.gender_listbox = ttk.Combobox(self.init_window_name, values=["Male", "Female", "Unable to determine"])
        self.gender_listbox.place(x=-500+1460,y=720)
        self.gender_listbox.config(font=(20))
        self.gender_listbox.bind("<<ComboboxSelected>>", self.gender_callback)

        self.gender_note = Label(self.init_window_name, text="note:")
        self.gender_note.place(x=-500+1680,y=720)
        self.gender_note.config(font=(20))

        self.gender_note_entry = Entry(self.init_window_name, font=(20), width=25)
        self.gender_note_entry.place(x=-500+1730,y=720)
        self.gender_note_entry.bind("<Return>", self.gender_note_entry_callback)

        # save
        self.save_button = Button(self.init_window_name, text="save", width=15, bg='light blue', command=self.save_coding)
        self.save_button.place(x=-500+1560,y=755)
        self.save_button.config(font=(30))

        

    def load_video(self, event=None):
        print("load video", self.event_num_entry.get())
        event_id = -1
        try:
            event_id = int(self.event_num_entry.get())
        except:
            return
        if event_id < 1 or event_id > 1260:
            return

        event_info = event_summary.iloc[event_id-1]
        self.event_id = event_id
        self.event_info = event_info
        #print("trip:", event_info[2], "start:", event_info[3], "end:", event_info[4])
        
        video_path = 'Videos/' + str(event_info['Event']) + '.mp4'
        # print(video_path)
        reader = cv2.VideoCapture(video_path)
        width = int(reader.get(cv2.CAP_PROP_FRAME_WIDTH))
        height = int(reader.get(cv2.CAP_PROP_FRAME_HEIGHT))
        self.fps = reader.get(cv2.CAP_PROP_FPS)
        frame_count = int(reader.get(cv2.CAP_PROP_FRAME_COUNT))

        self.start_frame = 0
        #self.end_frame = int((event_info[2] / 1000.0) * self.fps)
        reader.set(cv2.CAP_PROP_POS_FRAMES,0)
        #print(self.start_frame, self.end_frame)
        # print(reader.isOpened())

        have_more_frame = True
        curr_frame = self.start_frame
        self.saved_frames = []
        self.display_frame_id = 0
        while have_more_frame:
            have_more_frame, frame = reader.read()
            if curr_frame > self.start_frame:
                self.saved_frames.append(frame)
            curr_frame += 1

        print("length:", len(self.saved_frames))

        first_frame = self.saved_frames[0]
        first_frame = cv2.resize(first_frame, (720, 240))
        if first_frame.shape[0] > 480:
            first_frame = cv2.resize(first_frame, (1280, 720))
        self.img = ImageTk.PhotoImage(Image.fromarray(cv2.cvtColor(first_frame, cv2.COLOR_BGR2RGB)))
        self.canvas.create_image(0, 0, image=self.img, anchor=NW)
        
        # reinitialize things
        self.imcomplete_items = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10]
        self.update_imcomplete()

        self.right_of_way_note_entry.delete(0, END)
        self.secondary_activity_note_entry.delete(0, END)
        self.glance_note_entry.delete(0, END)
        self.hand_note_entry.delete(0, END)
        self.change_t2_note_entry.delete(0, END)
        self.reaction_t2_note_entry.delete(0, END)
        self.block_note_entry.delete(0, END)
        self.weather_note_entry.delete(0, END)
        self.age_group_note_entry.delete(0, END)
        self.gender_note_entry.delete(0, END)

        self.right_of_way_listbox.set('')
        self.hand_listbox.set('')
        self.reaction_t2_listbox.set('')
        self.block_listbox.set('')
        self.weather_listbox.set('')
        self.age_group_listbox.set('')
        self.gender_listbox.set('')

        self.checkvalue_1.set(False)
        self.checkvalue_2.set(False)
        self.checkvalue_3.set(False)
        self.checkvalue_4.set(False)
        self.checkvalue_5.set(False)
        self.checkvalue_6.set(False)
        self.checkvalue_7.set(False)
        self.checkvalue_8.set(False)
        self.checkvalue_9.set(False)
        self.checkvalue_10.set(False)
        self.checkvalue_11.set(False)
        self.checkvalue_12.set(False)

        # information of this event
        self.right_of_way_info = ("", "")
        self.secondary_activity_info = ([], "")
        self.glance_info = ("", "")
        self.hand_info = ("", "")
        self.change_t2_info = ("", "")
        self.reaction_t2_info = ("", "")
        self.block_info = ("", "")
        self.weather_info = ("", "")
        self.age_group_info = ("", "")
        self.gender_info = ("", "")
        

    def prev_frame(self, event=None):
        if self.display_frame_id > 0:
            self.display_frame_id -= 1
            p_frame = self.saved_frames[self.display_frame_id]
            p_frame = cv2.resize(p_frame, (720, 240))
            if p_frame.shape[0] > 480:
                p_frame = cv2.resize(p_frame, (1280, 720))
            self.img = ImageTk.PhotoImage(Image.fromarray(cv2.cvtColor(p_frame, cv2.COLOR_BGR2RGB)))
            self.canvas.create_image(0, 0, image=self.img, anchor=NW)

    def next_frame(self, event=None):
        if self.display_frame_id < len(self.saved_frames) - 1:
            self.display_frame_id += 1
            n_frame = self.saved_frames[self.display_frame_id]
            n_frame = cv2.resize(n_frame, (720, 240))
            if n_frame.shape[0] > 480:
                n_frame = cv2.resize(n_frame, (1280, 720))
            self.img = ImageTk.PhotoImage(Image.fromarray(cv2.cvtColor(n_frame, cv2.COLOR_BGR2RGB)))
            self.canvas.create_image(0, 0, image=self.img, anchor=NW)

    def prev_second(self, event=None):
        if self.display_frame_id - self.fps > 0:
            self.display_frame_id -= int(self.fps)
        else:
            self.display_frame_id = 0
        p_frame = self.saved_frames[self.display_frame_id]
        p_frame = cv2.resize(p_frame, (720, 240))
        if p_frame.shape[0] > 480:
            p_frame = cv2.resize(p_frame, (1280, 720))
        self.img = ImageTk.PhotoImage(Image.fromarray(cv2.cvtColor(p_frame, cv2.COLOR_BGR2RGB)))
        self.canvas.create_image(0, 0, image=self.img, anchor=NW)

    def next_second(self, event=None):
        if self.display_frame_id + self.fps < len(self.saved_frames) - 1:
            self.display_frame_id += int(self.fps)
        else:
            self.display_frame_id = len(self.saved_frames) - 1
        n_frame = self.saved_frames[self.display_frame_id]
        n_frame = cv2.resize(n_frame, (720, 240))
        if n_frame.shape[0] > 480:
            n_frame = cv2.resize(n_frame, (1280, 720))
        self.img = ImageTk.PhotoImage(Image.fromarray(cv2.cvtColor(n_frame, cv2.COLOR_BGR2RGB)))
        self.canvas.create_image(0, 0, image=self.img, anchor=NW)

    def update_imcomplete(self):
        self.complete.delete(1.0, END)
        res_string = ""
        for i in self.imcomplete_items:
            res_string += str(i) + " "
        self.complete.insert("1.0", "imcomplete: " + res_string)

    def save_coding(self):
        workbook_write = openpyxl.Workbook()
        worksheet_write = workbook_write.active
        worksheet_write.title = "Sheet1"

        worksheet_write.cell(1, 1, 'Right of way')
        worksheet_write.cell(1, 3, 'Secondary activity')
        worksheet_write.cell(1, 5, 'First road user glance to study vehicle')
        worksheet_write.cell(1, 7, 'Body languages')
        worksheet_write.cell(1, 9, "ChangeT2")
        worksheet_write.cell(1, 11, "ReactionT2")
        worksheet_write.cell(1, 13, "Block")
        worksheet_write.cell(1, 15, "Weather")
        worksheet_write.cell(1, 17, "Age group")
        worksheet_write.cell(1, 19, "Gender")
        for i in range(1, 11):
            worksheet_write.cell(1, 2 * i, 'note')

        worksheet_write.cell(2, 1, self.right_of_way_info[0])
        worksheet_write.cell(2, 2, self.right_of_way_info[1])
        secondary_activity_string = ""
        print(self.secondary_activity_info[0])
        for ele in self.secondary_activity_info[0]:
            secondary_activity_string += ele + ", "
        worksheet_write.cell(2, 3, secondary_activity_string)
        worksheet_write.cell(2, 4, self.secondary_activity_info[1])
        worksheet_write.cell(2, 5, self.glance_info[0])
        worksheet_write.cell(2, 6, self.glance_info[1])
        worksheet_write.cell(2, 7, self.hand_info[0])
        worksheet_write.cell(2, 8, self.hand_info[1])
        worksheet_write.cell(2, 9, self.change_t2_info[0])
        worksheet_write.cell(2, 10, self.change_t2_info[1])
        worksheet_write.cell(2, 11, self.reaction_t2_info[0])
        worksheet_write.cell(2, 12, self.reaction_t2_info[1])
        worksheet_write.cell(2, 13, self.block_info[0])
        worksheet_write.cell(2, 14, self.block_info[1])
        worksheet_write.cell(2, 15, self.weather_info[0])
        worksheet_write.cell(2, 16, self.weather_info[1])
        worksheet_write.cell(2, 17, self.age_group_info[0])
        worksheet_write.cell(2, 18, self.age_group_info[1])
        worksheet_write.cell(2, 19, self.gender_info[0])
        worksheet_write.cell(2, 20, self.gender_info[1])

        workbook_write.save(filename='event_' + str(self.event_id) + '.xlsx')
        print("save")

    def right_of_way_callback(self, event):
        # print("right_of_way_callback")
        print("right of way:", self.right_of_way_listbox.get())
        self.right_of_way_info = (self.right_of_way_listbox.get(), self.right_of_way_info[1])
        print("right of way info:", self.right_of_way_info)
        if self.right_of_way_info[0] != "Others":
            try:
                self.imcomplete_items.remove(1)
            except:
                print("already removed")
        else:
            if self.right_of_way_info[1] != "":
                try:
                    self.imcomplete_items.remove(1)
                except:
                    print("already removed")
        self.update_imcomplete()

    def right_of_way_note_entry_callback(self, event):
        # print("right_of_way_note_entry_callback")
        print("right of way note:", self.right_of_way_note_entry.get())
        self.right_of_way_info = (self.right_of_way_info[0], self.right_of_way_note_entry.get())
        print("right of way info", self.right_of_way_info)
        if self.right_of_way_info[0] == "Others":
            if self.right_of_way_info[1] != "":
                try:
                    self.imcomplete_items.remove(1)
                except:
                    print("already removed")
        else:
            print("already removed")
        self.update_imcomplete()

    def secondary_activity_callback(self, event):
        print("secondary_activity_callback")

    def secondary_activity_note_entry_callback(self, event):
        # print("secondary_activity_note_entry_callback")
        print("secondary activity note:", self.secondary_activity_note_entry.get())
        self.secondary_activity_info = (self.secondary_activity_info[0], self.secondary_activity_note_entry.get())
        print("secondary activity info", self.secondary_activity_info)
        if "Others" in self.secondary_activity_info[0] or "Unable to determine" in self.secondary_activity_info[0]:
            if self.secondary_activity_info[1] != "":
                try:
                    self.imcomplete_items.remove(2)
                except:
                    print("already removed")
        else:
            print("already removed")
        self.update_imcomplete()

    def glance_yes(self):
        self.glance_info = (self.display_frame_id, self.glance_info[1])
        try:
            self.imcomplete_items.remove(3)
        except:
            print("already removed")
        self.update_imcomplete()

    def glance_note_entry_callback(self, event):
        # print("glance_note_entry_callback")
        print("glance note:", self.glance_note_entry.get())
        self.glance_info = (self.glance_info[0], self.glance_note_entry.get())
        try:
            self.imcomplete_items.remove(3)
        except:
            print("already removed")
        self.update_imcomplete()

    def hand_callback(self, event):
        # print("hand_callback")
        print("body languages:", self.hand_listbox.get())
        self.hand_info = (self.hand_listbox.get(), self.hand_info[1])
        print("hand info:", self.hand_info)
        if self.hand_info[0] != "Others":
            try:
                self.imcomplete_items.remove(4)
            except:
                print("already removed")
        else:
            if self.hand_info[1] != "":
                try:
                    self.imcomplete_items.remove(4)
                except:
                    print("already removed")
        self.update_imcomplete()

    def hand_note_entry_callback(self, event):
        # print("hand_note_entry_callback")
        print("hand note:", self.hand_note_entry.get())
        self.hand_info = (self.hand_info[0], self.hand_note_entry.get())
        print("hand info", self.hand_info)
        if self.hand_info[0] == "Others":
            if self.hand_info[1] != "":
                try:
                    self.imcomplete_items.remove(4)
                except:
                    print("already removed")
        else:
            print("already removed")
        self.update_imcomplete()

    def change_t2_yes(self):
        # print("change_t2_yes")
        self.change_t2_info = (self.display_frame_id, self.change_t2_info[1])
        try:
            self.imcomplete_items.remove(5)
        except:
            print("already removed")
        self.update_imcomplete()

    def change_t2_note_entry_callback(self, event):
        # print("change_t2_note_entry_callback")
        print("change t2 note:", self.change_t2_note_entry.get())
        self.change_t2_info = (self.change_t2_info[0], self.change_t2_note_entry.get())
        try:
            self.imcomplete_items.remove(5)
        except:
            print("already removed")
        self.update_imcomplete()

    def reaction_t2_callback(self, event):
        # print("reaction_t2_callback")
        print("reaction t2:", self.reaction_t2_listbox.get())
        self.reaction_t2_info = (self.reaction_t2_listbox.get(), self.reaction_t2_info[1])
        print("reaction t2 info:", self.reaction_t2_info)
        if self.reaction_t2_info[0] != "Others" and self.reaction_t2_info[0] != "Not Applicable" and self.reaction_t2_info[0] != "Unable to determine":
            try:
                self.imcomplete_items.remove(6)
            except:
                print("already removed")
        else:
            if self.reaction_t2_info[1] != "":
                try:
                    self.imcomplete_items.remove(6)
                except:
                    print("already removed")
        self.update_imcomplete()

    def reaction_t2_note_entry_callback(self, event):
        # print("reaction_t2_note_entry_callback")
        print("reaction t2 note:", self.reaction_t2_note_entry.get())
        self.reaction_t2_info = (self.reaction_t2_info[0], self.reaction_t2_note_entry.get())
        print("reaction t2 info", self.reaction_t2_info)
        if self.reaction_t2_info[0] == "Others" or self.reaction_t2_info[0] == "Not Applicable" or self.reaction_t2_info[0] == "Unable to determine":
            if self.reaction_t2_info[1] != "":
                try:
                    self.imcomplete_items.remove(6)
                except:
                    print("already removed")
        else:
            print("already removed")
        self.update_imcomplete()

    def block_callback(self, event):
        # print("block_callback")
        print("block:", self.block_listbox.get())
        self.block_info = (self.block_listbox.get(), self.block_info[1])
        print("block info:", self.block_info)
        if self.block_info[0] == "No":
            try:
                self.imcomplete_items.remove(7)
            except:
                print("already removed")
        else:
            if self.block_info[1] != "":
                try:
                    self.imcomplete_items.remove(7)
                except:
                    print("already removed")
        self.update_imcomplete()

    def block_note_entry_callback(self, event):
        # print("block_note_entry_callback")
        print("block note:", self.block_note_entry.get())
        self.block_info = (self.block_info[0], self.block_note_entry.get())
        print("block info", self.block_info)
        if self.block_info[0] == "Yes" or self.block_info[0] == "Other":
            if self.block_info[1] != "":
                try:
                    self.imcomplete_items.remove(7)
                except:
                    print("already removed")
        else:
            print("already removed")
        self.update_imcomplete()

    def weather_callback(self, event):
        # print("weather_callback")
        print("weather:", self.weather_listbox.get())
        self.weather_info = (self.weather_listbox.get(), self.weather_info[1])
        print("weather info:", self.weather_info)
        if self.weather_info[0] != "Unable to determine":
            try:
                self.imcomplete_items.remove(8)
            except:
                print("already removed")
        else:
            if self.weather_info[1] != "":
                try:
                    self.imcomplete_items.remove(8)
                except:
                    print("already removed")
        self.update_imcomplete()
    
    def weather_note_entry_callback(self, event):
        # print("weather_note_entry_callback")
        print("weather note:", self.weather_note_entry.get())
        self.weather_info = (self.weather_info[0], self.weather_note_entry.get())
        print("weather info", self.weather_info)
        if self.weather_info[0] == "Unable to determine":
            if self.weather_info[1] != "":
                try:
                    self.imcomplete_items.remove(8)
                except:
                    print("already removed")
        else:
            print("already removed")
        self.update_imcomplete()

    def age_group_callback(self, event):
        # print("age_group_callback")
        print("age group:", self.age_group_listbox.get())
        self.age_group_info = (self.age_group_listbox.get(), self.weather_info[1])
        print("age group info:", self.age_group_info)
        if self.age_group_info[0] != "Unable to determine":
            try:
                self.imcomplete_items.remove(9)
            except:
                print("already removed")
        else:
            if self.age_group_info[1] != "":
                try:
                    self.imcomplete_items.remove(9)
                except:
                    print("already removed")
        self.update_imcomplete()

    def age_group_note_entry_callback(self, event):
        # print("age_group_note_entry_callback")
        print("age group note:", self.age_group_note_entry.get())
        self.age_group_info = (self.age_group_info[0], self.age_group_note_entry.get())
        print("age group info:", self.age_group_info)
        if self.age_group_info[0] == "Unable to determine":
            if self.age_group_info[1] != "":
                try:
                    self.imcomplete_items.remove(9)
                except:
                    print("already removed")
        else:
            print("already removed")
        self.update_imcomplete()

    def gender_callback(self, event):
        # print("gender_callback")
        print("gender:", self.gender_listbox.get())
        self.gender_info = (self.gender_listbox.get(), self.gender_info[1])
        print("gender info:", self.gender_info)
        if self.gender_info[0] != "Unable to determine":
            try:
                self.imcomplete_items.remove(10)
            except:
                print("already removed")
        else:
            if self.gender_info[1] != "":
                try:
                    self.imcomplete_items.remove(10)
                except:
                    print("already removed")
        self.update_imcomplete()

    def gender_note_entry_callback(self, event):
        # print("gender_note_entry_callback")
        print("gender note:", self.gender_note_entry.get())
        self.gender_info = (self.gender_info[0], self.gender_note_entry.get())
        print("gender info:", self.gender_info)
        if self.gender_info[0] == "Unable to determine":
            if self.gender_info[1] != "":
                try:
                    self.imcomplete_items.remove(10)
                except:
                    print("already removed")
        else:
            print("already removed")
        self.update_imcomplete()

    def checkbutton_append_helper(self, name):
        # print("checkbutton_append_helper")
        if not name in self.secondary_activity_info[0]:
            new_info = self.secondary_activity_info[0]
            new_info.append(name)
            self.secondary_activity_info = (new_info, self.secondary_activity_info[1])
        
    def checkbutton_remove_helper(self, name):
        if name in self.secondary_activity_info[0]:
            new_info = self.secondary_activity_info[0]
            new_info.remove(name)
            self.secondary_activity_info = (new_info, self.secondary_activity_info[1])

    def checkbutton_1_callback(self):
        print("checkbutton_1_callback")
        print(self.checkvalue_1.get())
        if self.checkvalue_1.get() == True:
            self.checkbutton_append_helper("None")
        else:
            self.checkbutton_remove_helper("None")
        try:
            self.imcomplete_items.remove(2)
        except:
            print("already removed")
        self.update_imcomplete()


    def checkbutton_2_callback(self):
        # print("checkbutton_2_callback")
        if self.checkvalue_2.get() == True:
            self.checkbutton_append_helper("Phone, listening or talking")
        else:
            self.checkbutton_remove_helper("Phone, listening or talking")
        try:
            self.imcomplete_items.remove(2)
        except:
            print("already removed")
        self.update_imcomplete()

    def checkbutton_3_callback(self):
        # print("checkbutton_3_callback")
        if self.checkvalue_3.get() == True:
            self.checkbutton_append_helper("Phone, looking and/or manipulating")
        else:
            self.checkbutton_remove_helper("Phone, looking and/or manipulating")
        try:
            self.imcomplete_items.remove(2)
        except:
            print("already removed")
        self.update_imcomplete()

    def checkbutton_4_callback(self):
        # print("checkbutton_4_callback")
        if self.checkvalue_4.get() == True:
            self.checkbutton_append_helper("Looking at personal item")
        else:
            self.checkbutton_remove_helper("Looking at personal item")
        try:
            self.imcomplete_items.remove(2)
        except:
            print("already removed")
        self.update_imcomplete()

    def checkbutton_5_callback(self):
        # print("checkbutton_5_callback")
        if self.checkvalue_5.get() == True:
            self.checkbutton_append_helper("Talking to someone in person")
        else:
            self.checkbutton_remove_helper("Talking to someone in person")
        try:
            self.imcomplete_items.remove(2)
        except:
            print("already removed")
        self.update_imcomplete()

    def checkbutton_6_callback(self):
        # print("checkbutton_6_callback")
        if self.checkvalue_6.get() == True:
            self.checkbutton_append_helper("Gesturing or interacting")
        else:
            self.checkbutton_remove_helper("Gesturing or interacting")
        try:
            self.imcomplete_items.remove(2)
        except:
            print("already removed")
        self.update_imcomplete()

    def checkbutton_7_callback(self):
        # print("checkbutton_7_callback")
        if self.checkvalue_7.get() == True:
            self.checkbutton_append_helper("Eating or drinking")
        else:
            self.checkbutton_remove_helper("Eating or drinking")
        try:
            self.imcomplete_items.remove(2)
        except:
            print("already removed")
        self.update_imcomplete()

    def checkbutton_8_callback(self):
        # print("checkbutton_8_callback")
        if self.checkvalue_8.get() == True:
            self.checkbutton_append_helper("Jogging, recreation")
        else:
            self.checkbutton_remove_helper("Jogging, recreation")
        try:
            self.imcomplete_items.remove(2)
        except:
            print("already removed")
        self.update_imcomplete()

    def checkbutton_9_callback(self):
        # print("checkbutton_9_callback")
        if self.checkvalue_9.get() == True:
            self.checkbutton_append_helper("Apparent rushing")
        else:
            self.checkbutton_remove_helper("Apparent rushing")
        try:
            self.imcomplete_items.remove(2)
        except:
            print("already removed")
        self.update_imcomplete()

    def checkbutton_10_callback(self):
        # print("checkbutton_10_callback")
        if self.checkvalue_10.get() == True:
            self.checkbutton_append_helper("Earphones")
        else:
            self.checkbutton_remove_helper("Earphones")
        try:
            self.imcomplete_items.remove(2)
        except:
            print("already removed")
        self.update_imcomplete()
    
    def checkbutton_11_callback(self):
        # print("checkbutton_11_callback")
        if self.checkvalue_11.get() == True:
            self.checkbutton_append_helper("Other")
        else:
            self.checkbutton_remove_helper("Other")
        if self.secondary_activity_info[1] != "":
            try:
                self.imcomplete_items.remove(2)
            except:
                print("already removed")
        self.update_imcomplete()
    
    def checkbutton_12_callback(self):
        # print("checkbutton_12_callback")
        if self.checkvalue_12.get() == True:
            self.checkbutton_append_helper("Unable to determine")
        else:
            self.checkbutton_remove_helper("Unable to determine")
        if self.secondary_activity_info[1] != "":
            try:
                self.imcomplete_items.remove(2)
            except:
                print("already removed")
        self.update_imcomplete()
    

def read_summary():
    event_summary = pd.read_csv("Events/Pedestrian/NCRC_Events-1.csv")

def main():
    read_summary()
    init_window = Tk()
    coding_GUI(init_window)
    init_window.mainloop()

if __name__ == '__main__':
    main()